from __future__ import annotations

import argparse
import csv
import json
import os
import random
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PROCESSED_DIR = ROOT / "data" / "processed"
DEFAULT_SOURCE_ROOT = ROOT / "data" / "source" / "raw" / "018.감성대화"
TRAIN_RELATIVE_PATH = Path(
    "Training_221115_add/원천데이터/감성대화말뭉치(최종데이터)_Training.xlsx"
)
VALIDATION_RELATIVE_PATH = Path(
    "Validation_221115_add/원천데이터/감성대화말뭉치(최종데이터)_Validation.xlsx"
)

LABELS = ["행복", "안정", "설렘", "중립", "불안", "피로", "우울", "화남"]
SOURCE_LABEL_TO_PROJECT_LABEL = {
    "기쁨": "행복",
    "슬픔": "우울",
    "분노": "화남",
    "불안": "불안",
    "상처": "우울",
    "당황": "불안",
}
RANDOM_SEED = 42
MAX_PER_LABEL_TRAIN = 2200
MAX_PER_LABEL_VALIDATION = 350

DIARY_ENDINGS = {
    "행복": [
        "덕분에 하루가 밝게 느껴졌고, 이 기분을 오래 기억하고 싶다.",
        "오늘을 떠올리면 웃음이 먼저 난다.",
        "내 노력이 의미 있게 느껴졌다.",
    ],
    "안정": [
        "마음이 차분해서 하루를 무리 없이 마무리할 수 있었다.",
        "큰 문제 없이 지나가서 안도감이 들었다.",
        "조용하고 평온한 기분이 이어졌다.",
    ],
    "설렘": [
        "앞으로 좋은 일이 생길 것 같아 마음이 두근거렸다.",
        "기대되는 마음 때문에 하루가 조금 들떴다.",
        "새로운 가능성이 보여서 기분 좋은 긴장감이 남았다.",
    ],
    "불안": [
        "앞으로 어떻게 될지 몰라 마음이 계속 불안했다.",
        "괜찮을 거라고 생각하려 해도 걱정이 쉽게 사라지지 않았다.",
        "머릿속이 복잡해서 편하게 쉬기 어려웠다.",
    ],
    "피로": [
        "몸과 마음이 모두 지쳐서 회복이 필요했다.",
        "에너지가 거의 남지 않은 느낌이었다.",
        "오늘은 무리하지 않고 쉬어야겠다고 느꼈다.",
    ],
    "우울": [
        "마음 한쪽이 무겁고 쉽게 털어지지 않았다.",
        "괜찮은 척했지만 계속 가라앉는 느낌이었다.",
        "오늘은 유난히 혼자라는 생각이 많이 들었다.",
    ],
    "화남": [
        "생각할수록 화가 올라와서 마음이 쉽게 가라앉지 않았다.",
        "별일 아닌 척하려 했지만 계속 짜증이 남았다.",
        "내가 왜 이런 기분을 참아야 하는지 답답했다.",
    ],
}

NEUTRAL_SITUATIONS = [
    "아침에 일어나 간단히 밥을 먹고 하루를 시작했다.",
    "오늘은 특별한 일 없이 평소처럼 시간이 흘러갔다.",
    "해야 할 일을 조금 정리하고 책상 주변을 치웠다.",
    "집에 돌아와 씻고 잠깐 쉬었다.",
    "날씨를 확인하고 내일 일정을 머릿속으로 정리했다.",
    "점심은 평소처럼 먹었고 오후에는 조용히 시간을 보냈다.",
    "오늘 하루는 크게 좋지도 나쁘지도 않은 편이었다.",
    "퇴근 후에 잠시 산책하고 집으로 돌아왔다.",
    "메시지를 몇 개 확인하고 필요한 답장만 했다.",
    "잠들기 전에 내일 할 일을 간단히 적어두었다.",
]

NEUTRAL_ENDINGS = [
    "마음은 대체로 차분했고 큰 감정 변화는 없었다.",
    "그냥 평범한 하루였다고 느꼈다.",
    "특별히 기억에 남는 감정은 없지만 나쁘지 않았다.",
    "조용하게 지나간 하루였다.",
]

SYNTHETIC_DIARY_SEEDS = {
    "행복": [
        "오늘은 생각보다 일이 잘 풀려서 기분이 좋았다",
        "작은 성취였지만 스스로가 꽤 자랑스러웠다",
        "기대하지 않았던 칭찬을 받아서 하루 종일 기분이 좋았다",
        "좋은 소식을 듣고 하루 종일 웃음이 났다",
    ],
    "안정": [
        "오랜만에 마음이 편안하고 차분했다",
        "큰 문제 없이 하루가 지나가서 안도감이 들었다",
        "조용히 산책하고 나니 마음이 안정됐다",
        "해야 할 일을 정리하고 나니 숨이 조금 놓였다",
    ],
    "설렘": [
        "새로운 일을 앞두고 기대감 때문에 마음이 두근거렸다",
        "내일 만날 생각을 하니 괜히 설레고 기분이 들떴다",
        "처음 해보는 도전이라 긴장되지만 기대가 더 컸다",
        "좋은 변화가 생길 것 같아서 하루 종일 마음이 들떴다",
    ],
    "불안": [
        "앞으로 일이 어떻게 될지 몰라 계속 걱정됐다",
        "별문제 없을 거라고 생각하려 해도 마음이 초조했다",
        "해야 할 일이 많은데 어디서부터 시작해야 할지 막막했다",
        "실수할까 봐 계속 긴장되고 마음이 불편했다",
        "아직 일어나지 않은 일까지 떠올라 잠이 잘 오지 않았다",
    ],
    "피로": [
        "하루 종일 몸이 무겁고 아무것도 하고 싶지 않았다",
        "계속되는 일 때문에 지치고 에너지가 바닥난 느낌이었다",
        "잠을 자도 피곤이 풀리지 않아 무기력했다",
        "해야 할 일은 많은데 몸과 마음이 따라주지 않았다",
    ],
    "우울": [
        "괜찮은 척했지만 마음이 계속 가라앉았다",
        "사람들 사이에 있어도 혼자인 것 같은 기분이 들었다",
        "별일 아닌데도 눈물이 날 것처럼 마음이 무거웠다",
        "후회와 실망이 계속 남아 하루가 길게 느껴졌다",
    ],
    "화남": [
        "사소한 말 하나에도 짜증이 올라와서 참기 어려웠다",
        "내 노력을 당연하게 여기는 것 같아 화가 났다",
        "계속 억울한 생각이 들어서 마음이 가라앉지 않았다",
        "상대의 태도가 너무 무례해서 하루 종일 신경질이 났다",
    ],
}

SYNTHETIC_DIARY_CLOSINGS = {
    "행복": ["이런 날이 자주 있었으면 좋겠다.", "오늘의 기분을 오래 기억하고 싶다."],
    "안정": ["오늘은 이 평온함을 지키고 싶다.", "차분하게 마무리할 수 있어서 다행이다."],
    "설렘": ["좋은 방향으로 이어졌으면 좋겠다.", "기대되는 마음을 잘 간직하고 싶다."],
    "불안": ["차분해지려고 해도 걱정이 쉽게 줄지 않았다.", "일단 할 수 있는 것부터 정리해야겠다."],
    "피로": ["오늘은 무리하지 않고 쉬어야겠다.", "잠깐 멈추고 회복할 시간이 필요하다."],
    "우울": ["내일은 조금이라도 나아졌으면 좋겠다.", "오늘은 그냥 버티는 것만으로도 힘들었다."],
    "화남": ["생각할수록 마음이 거칠어져서 정리가 필요했다.", "이 감정을 그냥 넘기기는 어려웠다."],
}


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = " ".join(text.split())
    return text


def iter_source_rows(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    next(rows, None)
    for row in rows:
        label = clean_text(row[5])
        sub_label = clean_text(row[6])
        situation = clean_text(row[3])
        human_parts = [clean_text(row[i]) for i in (7, 9, 11) if i < len(row)]
        human_parts = [part for part in human_parts if part]
        if label and human_parts:
            yield {
                "label": label,
                "sub_label": sub_label,
                "situation": situation,
                "human_text": " ".join(human_parts),
            }


def make_diary_text(row: dict[str, str], variant: int) -> str:
    label = row["label"]
    situation = row["situation"].replace(",", ", ")
    human_text = row["human_text"]
    ending = random.choice(DIARY_ENDINGS.get(label, ["오늘의 감정이 오래 남았다."]))

    templates = [
        "오늘은 {human_text} {ending}",
        "오늘 {situation}와 관련해서 이런 생각이 들었다. {human_text} {ending}",
        "하루를 정리해보면 {human_text} {ending}",
        "오늘 있었던 일을 떠올리면 {human_text} {ending}",
    ]
    template = templates[variant % len(templates)]
    return template.format(human_text=human_text, situation=situation or "일상", ending=ending)


def build_from_xlsx(path: Path, max_per_label: int, source_name: str) -> list[dict[str, str]]:
    random.seed(RANDOM_SEED)
    by_label: dict[str, list[dict[str, str]]] = {label: [] for label in LABELS if label != "중립"}

    for row in iter_source_rows(path):
        label = SOURCE_LABEL_TO_PROJECT_LABEL.get(row["label"])
        if label in by_label:
            mapped_row = dict(row)
            mapped_row["label"] = label
            by_label[label].append(mapped_row)

    records: list[dict[str, str]] = []
    for label, rows in by_label.items():
        random.shuffle(rows)
        for index, row in enumerate(rows[:max_per_label]):
            text = make_diary_text(row, index)
            records.append(
                {
                    "text": text,
                    "main_emotion": label,
                    "sub_emotion": row["sub_label"],
                    "situation": row["situation"],
                    "source": source_name,
                }
            )
    random.shuffle(records)
    return records


def build_neutral_records(count: int, source_name: str) -> list[dict[str, str]]:
    random.seed(RANDOM_SEED + count)
    records = []
    for index in range(count):
        situation = random.choice(NEUTRAL_SITUATIONS)
        ending = random.choice(NEUTRAL_ENDINGS)
        connector = random.choice(["", " 그리고 ", " 이후에는 "])
        extra = random.choice(NEUTRAL_SITUATIONS)
        text = f"{situation}{connector}{extra} {ending}" if connector else f"{situation} {ending}"
        records.append(
            {
                "text": text,
                "main_emotion": "중립",
                "sub_emotion": "평범한",
                "situation": "일상",
                "source": source_name,
            }
        )
    return records


def build_synthetic_emotion_records(count_per_label: int, source_name: str) -> list[dict[str, str]]:
    random.seed(RANDOM_SEED + count_per_label + 100)
    records = []
    for label, seeds in SYNTHETIC_DIARY_SEEDS.items():
        closings = SYNTHETIC_DIARY_CLOSINGS[label]
        for index in range(count_per_label):
            first = random.choice(seeds)
            second = random.choice(seeds)
            closing = random.choice(closings)
            if first == second:
                text = f"{first}. {closing}"
            else:
                text = f"{first}. {second}. {closing}"
            records.append(
                {
                    "text": text,
                    "main_emotion": label,
                    "sub_emotion": "합성",
                    "situation": "다이어리",
                    "source": source_name,
                }
            )
    return records


def write_csv(path: Path, records: list[dict[str, str]]) -> None:
    fieldnames = ["text", "main_emotion", "sub_emotion", "situation", "source"]
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="AI Hub 감성대화 말뭉치를 다이어리 감정분석 데이터셋으로 변환합니다."
    )
    parser.add_argument(
        "--source-root",
        default=os.environ.get("DIARY_EMOTION_SOURCE_ROOT", str(DEFAULT_SOURCE_ROOT)),
        help=(
            "018.감성대화 폴더 경로입니다. 생략하면 "
            "data/source/raw/018.감성대화 를 사용합니다."
        ),
    )
    return parser.parse_args()


def resolve_source_files(source_root: Path) -> tuple[Path, Path]:
    train_xlsx = source_root / TRAIN_RELATIVE_PATH
    validation_xlsx = source_root / VALIDATION_RELATIVE_PATH
    missing = [path for path in (train_xlsx, validation_xlsx) if not path.exists()]
    if missing:
        paths = "\n".join(f"- {path}" for path in missing)
        raise FileNotFoundError(
            "원본 엑셀 파일을 찾지 못했습니다.\n"
            "다음 둘 중 하나로 해결하세요.\n"
            "1. 원본 018.감성대화 폴더를 data/source/raw/ 아래에 복사\n"
            "2. --source-root 옵션으로 원본 폴더 경로 지정\n\n"
            f"없는 파일:\n{paths}"
        )
    return train_xlsx, validation_xlsx


def main() -> None:
    args = parse_args()
    train_xlsx, validation_xlsx = resolve_source_files(Path(args.source_root))
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)

    train_records = build_from_xlsx(train_xlsx, MAX_PER_LABEL_TRAIN, "aihub_diary_style_train")
    train_records.extend(build_synthetic_emotion_records(1000, "synthetic_diary_train"))
    train_records.extend(build_neutral_records(1000, "synthetic_neutral_train"))
    random.shuffle(train_records)

    validation_records = build_from_xlsx(
        validation_xlsx, MAX_PER_LABEL_VALIDATION, "aihub_diary_style_validation"
    )
    validation_records.extend(build_neutral_records(200, "synthetic_neutral_validation"))
    random.shuffle(validation_records)

    write_csv(PROCESSED_DIR / "diary_emotion_train.csv", train_records)
    write_csv(PROCESSED_DIR / "diary_emotion_validation.csv", validation_records)

    label_map = {
        "labels": LABELS,
        "description": {
            "행복": "즐거움, 만족, 감사, 성취감처럼 긍정적이고 밝은 정서",
            "안정": "편안함, 안도감, 차분함, 평온함처럼 마음이 안정된 정서",
            "설렘": "기대, 두근거림, 신남, 새로운 가능성에 대한 긍정적 긴장감",
            "중립": "큰 감정 변화 없이 평범하고 차분한 상태",
            "불안": "걱정, 두려움, 초조, 긴장처럼 미래 위험을 예상하는 정서",
            "피로": "지침, 무기력, 번아웃, 체력 저하처럼 에너지가 고갈된 상태",
            "우울": "슬픔, 외로움, 실망, 후회처럼 가라앉고 무거운 정서",
            "화남": "화, 짜증, 억울함, 답답함처럼 격한 불쾌감이 올라온 정서",
        },
    }
    (PROCESSED_DIR / "label_map.json").write_text(
        json.dumps(label_map, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    summary = {
        "train_total": len(train_records),
        "validation_total": len(validation_records),
        "train_distribution": Counter(r["main_emotion"] for r in train_records),
        "validation_distribution": Counter(r["main_emotion"] for r in validation_records),
        "notes": [
            "AI Hub 감성대화 사람문장을 다이어리 문체로 변환했습니다.",
            "프로젝트 감정 대분류를 행복, 안정, 설렘, 중립, 불안, 피로, 우울, 화남 8개로 구성했습니다.",
            "원본에 부족한 라벨은 합성 다이어리 문장으로 보강했습니다.",
            "실제 사용자 다이어리와 문체 차이가 있으므로 서비스 적용 전 샘플 검수가 필요합니다.",
        ],
    }
    (PROCESSED_DIR / "dataset_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
